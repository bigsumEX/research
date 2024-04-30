import pandas as pd

from openpyxl import Workbook
from openpyxl.drawing.image import Image  # ここを修正
from function.SEM import run_SEM
import semopy
import os


def sem_plot(obj, num=10, filename="./output/sem/output.xlsx"):
    df = obj.study.trials_dataframe()
    df = df[["number", "value", "params_beta", "params_w_threshold", "user_attrs_best_sm"]]
    df.columns = ["number", "result", "beta", "w_threshold", "user_attrs"]
    df = pd.concat([df, obj.df_stats.reset_index(drop=True)], axis=1)

    df = df.drop('Value', axis=1)
    df = df.dropna()

    #昇順の場合：ascending=False
    #インデックスを振り直す: ignore_index=True
    df_top = df.sort_values("result",ignore_index=True, ascending=False).head(num)
    df_top = df_top.drop("user_attrs", axis=1)

    df_result = df_top.loc[:, :"w_threshold"]
    df_stats = df_top.loc[:, "DoF":]
    print(df_top)

    # Excelファイルを作成
    wb = Workbook()
    ws = wb.active
    ws.title = "Top Results"

    # カラム名を書き込む
    ws.append(list(df_result.columns))
    # DataFrameをExcelに書き込む
    for r_idx, row in enumerate(df_result.itertuples(), start=2):
        for c_idx, value in enumerate(row[1:], start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # カラム名を書き込む
    ws.append(list(df_stats.columns))
    # DataFrameをExcelに書き込む
    for r_idx, row in enumerate(df_stats.itertuples(), start=len(df_top) + 4):
        for c_idx, value in enumerate(row[1:], start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)


    for row in df_top.iterrows():
         # 新しいシートを作成し、各行のデータフレームを書き込む
        ws_new = wb.create_sheet(title="SEM Result {}".format(int(row[1]['number'])))
        ws_new.append(list(df_top.columns))

        row_data = [row[1][col] for col in df_top.columns]
        ws_new.append(row_data)
        sm_SEM = run_SEM(obj.df, obj.matrix_list[int(row[1]["number"])], row[1]["w_threshold"])       
        
        img_path = f"./output/sem/{int(row[1]['number'])}_semopy.png"
        semopy.semplot(sm_SEM[0], img_path,
                                    engine="dot",        # 階層的なグラフを生成するエンジン(デフォルト)
                                    plot_covs=True,      # Ture: 共分散がプロット
                                    std_ests=True)       # Ture: 標準化された推定値をプロット
        
        # 画像を対応するシートに挿入します
        img = Image(img_path)  # ここを修正
        img.anchor = ws_new.cell(row=3, column=1).coordinate
        ws_new.add_image(img)

    wb.save(filename)

    # # 後処理：一時ファイルを削除します
    # for row in df_top.iterrows():
    #     img_path = f"./output/sem/{int(row[1]['number'])}_semopy.png"
    #     if os.path.exists(img_path):
    #         os.remove(img_path)
